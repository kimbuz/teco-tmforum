"""
DTW Ignite 2026 — Export calendar data to Excel (.xlsx)
Reads the enriched JSON files and creates a formatted Excel workbook.
"""

import json
import os

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl...")
    os.system("pip install openpyxl")
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

RAW_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "raw")
OUTPUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "DTW_Ignite_2026_Agenda.xlsx")

# Column config
COLUMNS = [
    ("Day", 12),
    ("Time", 14),
    ("Title", 55),
    ("Type", 18),
    ("Stage", 30),
    ("Track", 30),
    ("Access", 22),
    ("Speakers", 50),
    ("Description", 80),
    ("URL", 40),
]

# Colors per day
DAY_COLORS = {
    "Tuesday 23": "DCE6F1",    # Light blue
    "Wednesday 24": "E2EFDA",  # Light green
    "Thursday 25": "FFF2CC",   # Light yellow
}


def self_detect_type(title):
    """Detect session type from title prefix."""
    TYPES = ["Keynote", "Catalyst", "Spotlight", "Awards", "Masterclass"]
    for t in TYPES:
        if title.startswith(t):
            return t
    # Also check for common patterns
    if "Roundtable" in title:
        return "Roundtable"
    if "Lunch briefing" in title or "Breakfast briefing" in title:
        return "Briefing"
    if "The Loft" in title:
        return "Loft"
    if "Mission Garage" in title:
        return "Mission Garage"
    if "Networking break" in title or "Networking lunch" in title:
        return "Networking"
    if "Circle" in title:
        return "Circle"
    return ""


def load_sessions():
    """Load all sessions from JSON files."""
    day_files = [
        ("Tuesday 23", "day1_tuesday_23.json"),
        ("Wednesday 24", "day2_wednesday_24.json"),
        ("Thursday 25", "day3_thursday_25.json"),
    ]

    all_sessions = []
    for day_name, filename in day_files:
        filepath = os.path.join(RAW_DIR, filename)
        with open(filepath, "r", encoding="utf-8") as f:
            sessions = json.load(f)

        for s in sessions:
            detail = s.get("detail", {})
            speakers_list = detail.get("speakers", [])
            speakers_str = ", ".join(
                f"{sp['name']} ({sp['role']})" for sp in speakers_list if sp.get("name")
            )

            tracks = detail.get("tracks", [])
            tracks_str = ", ".join(tracks) if tracks else ""

            # Determine access level
            desc = detail.get("description", "")
            title = detail.get("full_title", "")
            raw_text = " ".join(s.get("raw_lines", []))
            all_text = (desc + " " + title + " " + raw_text).lower()

            access = "Open"
            if "private" in all_text and ("circle" in all_text or "csp" in all_text):
                access = "Invite Only (CSP executives)"
            elif "by invitation" in all_text and "leadership forum" in all_text:
                access = "Invite Only (Leadership Forum)"
            elif "by invitation" in all_text and ("lunch" in all_text or "breakfast" in all_text):
                access = "RSVP (login to member account)"
            elif "by invitation" in all_text and "roundtable" in all_text:
                access = "RSVP (login to member account)"
            elif "by invitation" in all_text and "spotlight" in all_text:
                access = "RSVP (login to member account)"
            elif "by invitation" in all_text and "decode" in all_text:
                access = "RSVP (login to member account)"
            elif "by invitation" in all_text:
                access = "RSVP (login to member account)"
            elif "leadership forum" in all_text:
                access = "Invite Only (Leadership Forum)"
            elif "vip" in all_text or "exclusive" in all_text:
                access = "VIP pass required"

            all_sessions.append({
                "day": day_name,
                "time": detail.get("time_slot", ""),
                "title": detail.get("full_title", s.get("title", "")),
                "type": detail.get("session_type", s.get("type", "")) or self_detect_type(detail.get("full_title", "")),
                "stage": detail.get("stage", s.get("stage", "")),
                "track": tracks_str,
                "access": access,
                "speakers": speakers_str,
                "description": detail.get("description", s.get("description", ""))[:500],
                "url": s.get("url", ""),
            })

    return all_sessions


def create_excel(sessions):
    """Create formatted Excel workbook."""
    wb = Workbook()
    ws = wb.active
    ws.title = "DTW Ignite 2026 Agenda"

    # Styles
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Write headers
    for col_idx, (col_name, col_width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Write data
    for row_idx, session in enumerate(sessions, 2):
        values = [
            session["day"],
            session["time"],
            session["title"],
            session["type"],
            session["stage"],
            session["track"],
            session["access"],
            session["speakers"],
            session["description"],
            session["url"],
        ]

        # Day color
        day_color = DAY_COLORS.get(session["day"], "FFFFFF")
        row_fill = PatternFill(start_color=day_color, end_color=day_color, fill_type="solid")

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = cell_align
            cell.border = thin_border
            cell.fill = row_fill

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(sessions)+1}"

    # Set row height
    for row in range(2, len(sessions) + 2):
        ws.row_dimensions[row].height = 30

    ws.row_dimensions[1].height = 20

    # Save
    wb.save(OUTPUT)
    print(f"Excel saved: {OUTPUT}")
    print(f"Total rows: {len(sessions)}")


def main():
    print("DTW Ignite 2026 — Export to Excel")
    print("=" * 50)

    sessions = load_sessions()
    print(f"Loaded {len(sessions)} sessions")

    create_excel(sessions)
    print("Done!")


if __name__ == "__main__":
    main()
